# Supporting script
# Write list of receipt data to the Excel spreadsheet

# Improvements:
#   - Check if something is already in the cell it's writing to
#   - Find a way to read an Excel sheet without erasing embedded cell equations
#   - Figure out how to prevent openpyxl from eliminating merged cell borders when writing

from openpyxl import load_workbook  # Workbook, load_workbook


def write(total_list, data_col, date_list, filename):
    input_row = 4
    counter = 0

    wb = load_workbook(filename)

    ws = wb["Expenses"]

    for i in total_list:
        while True:
            if date_list[counter] == i[0]:
                ws[data_col[0] + str(counter + input_row)] = i[1]
                ws[data_col[1] + str(counter + input_row)] = i[2]
                counter += 1
                break
            else:
                counter += 1

    wb.save(filename)
